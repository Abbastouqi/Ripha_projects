"""
FastAPI entry point.

Endpoints
─────────
GET  /stream            MJPEG live camera feed with face annotations
GET  /health            Service health + enrolled face count

POST /enroll            Enroll a new person (multipart: name, employee_id, images)
GET  /attendance        Query attendance records (optional ?date=YYYY-MM-DD)
GET  /unknown-faces     List unreviewed unknown faces
PATCH /unknown-faces/{id}/review  Mark unknown face reviewed

WS   /ws/events         Real-time attendance + unknown-face events
"""

from __future__ import annotations

import asyncio
import io
import sys
from contextlib import asynccontextmanager
from typing import Optional

# Prevents WinError 64 crash on Windows when network changes occur
if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

import cv2
import numpy as np
from fastapi import (
    FastAPI, File, Form, HTTPException, UploadFile, WebSocket,
    WebSocketDisconnect,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from supabase import create_client

from .config import (
    SUPABASE_URL, SUPABASE_KEY, INSIGHTFACE_MODEL,
    SIMILARITY_THRESHOLD, CAMERA_ID, FRAME_SKIP,
)
from .face_engine import FaceEngine
from .tracker import FaceTracker
from .matcher import FaceMatcher
from .attendance import AttendanceService
from .camera import CameraProcessor


# --------------------------------------------------------------------------- #
#  Application state (populated during lifespan)                              #
# --------------------------------------------------------------------------- #
_processor: Optional[CameraProcessor] = None
_matcher: Optional[FaceMatcher] = None
_engine: Optional[FaceEngine] = None
_attendance: Optional[AttendanceService] = None


@asynccontextmanager
async def lifespan(app: FastAPI):
    global _processor, _matcher, _engine, _attendance

    if not SUPABASE_URL or not SUPABASE_KEY:
        raise RuntimeError("SUPABASE_URL and SUPABASE_KEY must be set in .env")

    db = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Load enrolled embeddings into FAISS (graceful if tables not created yet)
    _matcher = FaceMatcher(threshold=SIMILARITY_THRESHOLD)
    try:
        result = (
            db.table("embeddings")
            .select("person_id, persons(name), embedding")
            .execute()
        )
        rows = [
            {
                "person_id": r["person_id"],
                "name": r["persons"]["name"],
                "embedding": r["embedding"],
            }
            for r in result.data
            if r.get("persons")
        ]
        _matcher.load(rows)
        print(f"[startup] Loaded {_matcher.enrolled_count} face embeddings from Supabase")
    except Exception as exc:
        print(f"[startup] WARNING: Could not load embeddings ({exc})")
        print("[startup] Run supabase_schema.sql in Supabase SQL Editor, then POST /reload-index")

    _engine = FaceEngine(model_name=INSIGHTFACE_MODEL)
    print(f"[startup] FaceEngine ready ({INSIGHTFACE_MODEL})")

    _attendance = AttendanceService(db)

    _processor = CameraProcessor(
        engine=_engine,
        tracker=FaceTracker(),
        matcher=_matcher,
        attendance=_attendance,
        camera_id=CAMERA_ID,
        frame_skip=FRAME_SKIP,
    )
    _processor.start(asyncio.get_event_loop())

    yield  # ── app is running ──

    _processor.stop()


app = FastAPI(title="SmartAttendance", version="1.0.0", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# --------------------------------------------------------------------------- #
#  Camera stream                                                               #
# --------------------------------------------------------------------------- #
@app.get("/stream", summary="MJPEG live feed")
async def stream():
    async def generate():
        while True:
            frame = _processor.latest_frame()
            if frame is None:
                await asyncio.sleep(0.03)
                continue
            _, jpeg = cv2.imencode(
                ".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 75]
            )
            yield (
                b"--frame\r\nContent-Type: image/jpeg\r\n\r\n"
                + jpeg.tobytes()
                + b"\r\n"
            )
            await asyncio.sleep(0.033)

    return StreamingResponse(
        generate(),
        media_type="multipart/x-mixed-replace; boundary=frame",
    )


@app.get("/snapshot", summary="Single JPEG frame (for lightweight clients)")
async def snapshot():
    from fastapi.responses import Response as RawResponse
    frame = _processor.latest_frame()
    if frame is None:
        raise HTTPException(503, "No frame available yet")
    _, jpeg = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 60])
    return RawResponse(content=jpeg.tobytes(), media_type="image/jpeg")


# --------------------------------------------------------------------------- #
#  Health                                                                      #
# --------------------------------------------------------------------------- #
@app.get("/health")
async def health():
    return {
        "status": "ok",
        "enrolled_faces": _matcher.enrolled_count if _matcher else 0,
        "model": INSIGHTFACE_MODEL,
    }


# --------------------------------------------------------------------------- #
#  Enroll                                                                      #
# --------------------------------------------------------------------------- #
@app.post("/enroll", summary="Enroll a new person")
async def enroll(
    name: str = Form(...),
    employee_id: str = Form(...),
    email: Optional[str] = Form(None),
    department: Optional[str] = Form(None),
    images: list[UploadFile] = File(..., description="3–5 clear face photos"),
):
    from supabase import create_client

    db = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Insert person record
    try:
        person_row = (
            db.table("persons")
            .insert(
                {
                    "name": name,
                    "employee_id": employee_id,
                    "email": email,
                    "department": department,
                }
            )
            .execute()
        )
        person_id = person_row.data[0]["id"]
    except Exception as exc:
        raise HTTPException(400, f"Could not create person record: {exc}")

    # Extract and store embeddings
    added = 0
    engine = FaceEngine(model_name=INSIGHTFACE_MODEL)
    for upload in images:
        data = await upload.read()
        arr = np.frombuffer(data, dtype=np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        if img is None:
            continue

        faces = engine.process(img)
        if not faces:
            continue

        # Use the largest face if multiple detected
        face = max(faces, key=lambda f: (f.bbox[2] - f.bbox[0]) * (f.bbox[3] - f.bbox[1]))
        embedding: np.ndarray = face.normed_embedding

        db.table("embeddings").insert(
            {"person_id": person_id, "embedding": embedding.tolist()}
        ).execute()

        # Add to live FAISS index immediately
        _matcher.add(person_id, name, embedding)
        added += 1

    if added == 0:
        db.table("persons").delete().eq("id", person_id).execute()
        raise HTTPException(
            400,
            "No faces detected in any uploaded image. "
            "Use clear, front-facing photos with good lighting.",
        )

    return {"person_id": person_id, "name": name, "embeddings_added": added}


# --------------------------------------------------------------------------- #
#  Presence — who is currently IN                                              #
# --------------------------------------------------------------------------- #
@app.get("/presence", summary="Who is currently inside")
async def get_presence():
    return _attendance.get_presence_snapshot()


@app.get("/presence/log", summary="Check-in / check-out event history")
async def get_presence_log(date: Optional[str] = None, limit: int = 100):
    db = create_client(SUPABASE_URL, SUPABASE_KEY)
    query = (
        db.table("presence_log")
        .select("*, persons(name, employee_id)")
        .order("timestamp", desc=True)
        .limit(limit)
    )
    if date:
        query = query.gte("timestamp", f"{date}T00:00:00+00:00").lt("timestamp", f"{date}T23:59:59+00:00")
    return query.execute().data


# --------------------------------------------------------------------------- #
#  Attendance records                                                          #
# --------------------------------------------------------------------------- #
@app.get("/attendance")
async def get_attendance(date: Optional[str] = None, limit: int = 100):
    db = create_client(SUPABASE_URL, SUPABASE_KEY)
    query = (
        db.table("presence_log")
        .select("*, persons(name, employee_id, department)")
        .eq("event_type", "checkin")
        .order("timestamp", desc=True)
        .limit(limit)
    )
    if date:
        query = query.gte("timestamp", f"{date}T00:00:00+00:00").lt("timestamp", f"{date}T23:59:59+00:00")
    return query.execute().data


# --------------------------------------------------------------------------- #
#  Unknown faces                                                               #
# --------------------------------------------------------------------------- #
@app.get("/unknown-faces")
async def list_unknown(limit: int = 50):
    db = create_client(SUPABASE_URL, SUPABASE_KEY)
    return (
        db.table("unknown_faces")
        .select("id, timestamp, reviewed, reviewer_notes")
        .eq("reviewed", False)
        .order("timestamp", desc=True)
        .limit(limit)
        .execute()
        .data
    )


@app.get("/export/attendance", summary="Download attendance as Excel")
async def export_attendance(date: Optional[str] = None):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse as SR

    db = create_client(SUPABASE_URL, SUPABASE_KEY)
    query = (
        db.table("presence_log")
        .select("*, persons(name, employee_id, department)")
        .order("timestamp", desc=False)
    )
    if date:
        query = query.gte("timestamp", f"{date}T00:00:00+00:00").lt(
            "timestamp", f"{date}T23:59:59+00:00"
        )
    rows = query.execute().data

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Header style
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF")
    headers = ["#", "Name", "Employee ID", "Department", "Event", "Date", "Time", "Confidence", "Camera"]
    col_widths = [5, 22, 14, 18, 12, 14, 12, 14, 12]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # Alternating row colours
    fill_in  = PatternFill("solid", fgColor="D4EDDA")   # green tint for checkin
    fill_out = PatternFill("solid", fgColor="F8D7DA")   # red tint for checkout
    fill_alt = PatternFill("solid", fgColor="F2F2F2")

    for i, r in enumerate(rows, 1):
        p = r.get("persons") or {}
        ts = r.get("timestamp", "")[:19].replace("T", " ")
        date_part = ts[:10] if ts else ""
        time_part = ts[11:] if len(ts) > 10 else ""
        event = r.get("event_type", "")
        conf = r.get("confidence")
        conf_str = f"{conf*100:.1f}%" if conf else "—"
        fill = fill_in if event == "checkin" else fill_out if event == "checkout" else fill_alt
        row_data = [
            i, p.get("name", ""), p.get("employee_id", ""),
            p.get("department", ""), event.upper(),
            date_part, time_part, conf_str, r.get("camera_id", "main"),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center" if col not in (2, 3, 4) else "left")

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"attendance_{date or 'all'}.xlsx"
    return SR(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.patch("/unknown-faces/{face_id}/review")
async def review_unknown(face_id: str, notes: Optional[str] = None):
    db = create_client(SUPABASE_URL, SUPABASE_KEY)
    db.table("unknown_faces").update(
        {"reviewed": True, "reviewer_notes": notes}
    ).eq("id", face_id).execute()
    return {"reviewed": True}


# --------------------------------------------------------------------------- #
#  WebSocket — real-time attendance events                                     #
# --------------------------------------------------------------------------- #
@app.websocket("/ws/events")
async def ws_events(websocket: WebSocket):
    """
    Streams JSON events:
      {"type": "attendance", "person_name": "...", "confidence": 0.52, ...}
      {"type": "unknown_face"}
    """
    await websocket.accept()
    q = _attendance.subscribe()
    try:
        while True:
            event = await asyncio.wait_for(q.get(), timeout=30.0)
            await websocket.send_json(event)
    except (WebSocketDisconnect, asyncio.TimeoutError):
        pass
    finally:
        _attendance.unsubscribe(q)


# --------------------------------------------------------------------------- #
#  Reload FAISS index (call after bulk enrollment)                            #
# --------------------------------------------------------------------------- #
@app.post("/reload-index", summary="Reload FAISS from Supabase")
async def reload_index():
    db = create_client(SUPABASE_URL, SUPABASE_KEY)
    result = (
        db.table("embeddings")
        .select("person_id, persons(name), embedding")
        .execute()
    )
    rows = [
        {
            "person_id": r["person_id"],
            "name": r["persons"]["name"],
            "embedding": r["embedding"],
        }
        for r in result.data
        if r.get("persons")
    ]
    global _matcher
    _matcher = FaceMatcher(threshold=SIMILARITY_THRESHOLD)
    _matcher.load(rows)
    _processor.matcher = _matcher
    return {"enrolled_count": _matcher.enrolled_count}
